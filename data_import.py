"""data_import.py — 既存POSからのデータ移行ウィザード

2つの方法でデータを取り込める:
  [A] CSV/Excel ファイル (1種類ずつ)
  [B] SQLiteデータベースファイル (.db 一発で全種類)

列名は日本語/英語どちらでも自動マッチング。
"""
import io
import csv
import sqlite3
import tempfile
import os
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, HTTPException, Header, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from db_shared import SessionLocal, require_role

router = APIRouter(tags=["import"])
ADMIN_ROLES = ["owner", "manager"]

# ─────────────────────────── データ種別定義 ───────────────────────────

DATA_TYPES = {
    "casts": {
        "label": "キャスト",
        "icon": "👤",
        "fields": [
            {"key": "name", "label": "氏名 (必須)", "required": True, "aliases": ["name", "氏名", "名前", "なまえ", "キャスト名", "キャスト", "源氏名"]},
            {"key": "rank", "label": "ランク", "required": False, "default": "", "aliases": ["rank", "ランク", "等級", "ランキング"]},
            {"key": "hourly_rate", "label": "時給", "required": False, "type": "float", "default": 0, "aliases": ["hourly_rate", "時給", "じきゅう", "時間給"]},
            {"key": "drink_back_rate", "label": "ドリンクバック率", "required": False, "type": "rate", "default": 0, "aliases": ["drink_back_rate", "ドリンクバック率", "DB率", "ドリンク率", "ドリンクバック"]},
            {"key": "nom_fee_hon", "label": "本指名バック", "required": False, "type": "float", "default": 0, "aliases": ["nom_fee_hon", "本指名バック", "本指名", "本バック"]},
            {"key": "nom_fee_jyonai", "label": "場内指名バック", "required": False, "type": "float", "default": 0, "aliases": ["nom_fee_jyonai", "場内バック", "場内", "場内指名"]},
            {"key": "nom_fee_dohan", "label": "同伴バック", "required": False, "type": "float", "default": 0, "aliases": ["nom_fee_dohan", "同伴バック", "同伴"]},
            {"key": "floor_rate", "label": "場内固定額", "required": False, "type": "float", "default": 0, "aliases": ["floor_rate", "場内固定額", "場内固定"]},
        ],
    },
    "tables": {
        "label": "テーブル",
        "icon": "🪑",
        "fields": [
            {"key": "name", "label": "テーブル名 (必須)", "required": True, "aliases": ["name", "テーブル名", "卓名", "席名", "テーブル", "卓番"]},
        ],
    },
    "items": {
        "label": "メニュー",
        "icon": "🍸",
        "fields": [
            {"key": "name", "label": "商品名 (必須)", "required": True, "aliases": ["name", "商品名", "メニュー名", "商品", "メニュー", "ドリンク名"]},
            {"key": "category", "label": "カテゴリ (set/drink/bottle/food/other)", "required": False, "default": "drink", "aliases": ["category", "カテゴリ", "種別", "区分", "分類"]},
            {"key": "price", "label": "価格 (必須)", "required": True, "type": "float", "aliases": ["price", "価格", "金額", "値段", "単価"]},
            {"key": "stock", "label": "在庫数", "required": False, "type": "int", "default": 0, "aliases": ["stock", "在庫", "在庫数"]},
        ],
    },
    "customers": {
        "label": "顧客",
        "icon": "🤝",
        "fields": [
            {"key": "nickname", "label": "ニックネーム (必須)", "required": True, "aliases": ["nickname", "氏名", "名前", "ニックネーム", "顧客名", "お客様名", "お客様"]},
            {"key": "phone", "label": "電話番号", "required": False, "default": "", "aliases": ["phone", "電話", "電話番号", "TEL", "tel", "携帯"]},
            {"key": "memo", "label": "メモ", "required": False, "default": "", "aliases": ["memo", "メモ", "備考", "コメント", "ノート"]},
        ],
    },
}

# カテゴリ名の正規化（日本語→英語キー）
CATEGORY_MAP = {
    "セット": "set", "set": "set", "セット料金": "set",
    "ドリンク": "drink", "drink": "drink", "drinks": "drink",
    "ボトル": "bottle", "bottle": "bottle",
    "フード": "food", "food": "food", "料理": "food", "おつまみ": "food",
    "その他": "other", "other": "other",
}

# ─────────────────────────── ヘルパー ───────────────────────────

def detect_encoding(raw: bytes) -> str:
    """ファイルのエンコーディングを自動判定"""
    if raw.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    try:
        raw.decode('utf-8')
        return 'utf-8'
    except UnicodeDecodeError:
        return 'cp932'  # 日本Windowsデフォルト

def normalize_header(s: str) -> str:
    """ヘッダー文字列を正規化（空白除去・小文字化）"""
    return s.strip().lower().replace(' ', '').replace('　', '')

def find_field(headers: List[str], aliases: List[str]) -> Optional[int]:
    """エイリアスのいずれかに一致する列のインデックスを返す"""
    nheaders = [normalize_header(h) for h in headers]
    for a in aliases:
        na = normalize_header(a)
        for i, h in enumerate(nheaders):
            if h == na:
                return i
    return None

def cast_value(v: str, ftype: str, default=None):
    """型変換（float/int/rate等）"""
    if v is None:
        return default
    v = str(v).strip()
    if not v:
        return default
    try:
        if ftype == "float":
            # ¥や,を除去
            v = v.replace(',', '').replace('¥', '').replace('円', '').strip()
            return float(v) if v else default
        elif ftype == "int":
            v = v.replace(',', '').strip()
            return int(float(v)) if v else default
        elif ftype == "rate":
            # 30% や 0.3 のどちらでも対応
            v = v.replace('%', '').replace(',', '').strip()
            f = float(v) if v else 0
            return f / 100.0 if f > 1 else f
        else:
            return v
    except (ValueError, TypeError):
        return default

def parse_file(raw: bytes, filename: str) -> List[Dict[str, str]]:
    """CSV/Excel/TSV を自動判定してパース、{header: value} の辞書配列を返す"""
    fn = filename.lower()
    if fn.endswith(('.xlsx', '.xls')):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(400, "Excel対応にはopenpyxlが必要です")
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue  # 空行スキップ
            result.append({headers[i]: ("" if c is None else str(c)) for i, c in enumerate(row) if i < len(headers)})
        return result
    else:
        # CSV/TSV
        enc = detect_encoding(raw)
        text = raw.decode(enc, errors='replace')
        # 区切り文字を自動判定
        delim = '\t' if text.count('\t') > text.count(',') else ','
        reader = csv.DictReader(io.StringIO(text), delimiter=delim)
        return [{k: (v or "") for k, v in row.items() if k} for row in reader]

def auto_map(rows: List[Dict[str, str]], data_type: str) -> Dict[str, Any]:
    """ファイル内容を data_type のスキーマに自動マッピング
    返り値: {mapped: [...], errors: [...], column_mapping: {...}}
    """
    if not rows:
        return {"mapped": [], "errors": ["ファイルが空です"], "column_mapping": {}}
    spec = DATA_TYPES[data_type]
    headers = list(rows[0].keys())

    # 列マッピング
    mapping = {}  # field_key -> header
    for f in spec["fields"]:
        idx = find_field(headers, f["aliases"])
        if idx is not None:
            mapping[f["key"]] = headers[idx]

    # 必須列の存在チェック
    errors = []
    for f in spec["fields"]:
        if f.get("required") and f["key"] not in mapping:
            errors.append(f'必須項目「{f["label"]}」に該当する列が見つかりません。'
                          f'（受け付ける列名: {", ".join(f["aliases"][:3])} 等）')

    # 行ごとに変換
    mapped = []
    for ri, row in enumerate(rows, start=2):  # 2行目から（1行目はヘッダー）
        rec = {}
        skip = False
        for f in spec["fields"]:
            key = f["key"]
            raw_val = row.get(mapping.get(key, ''), '') if key in mapping else ''
            ftype = f.get("type", "str")
            default = f.get("default", "")
            val = cast_value(raw_val, ftype, default)
            if f.get("required") and (val is None or val == "" or val == default and default == ""):
                # 必須項目が空
                if not raw_val:
                    skip = True  # 必須空欄行はスキップ
                    break
            # カテゴリ正規化
            if key == "category" and isinstance(val, str):
                v = val.strip().lower()
                val = CATEGORY_MAP.get(v, CATEGORY_MAP.get(val.strip(), "drink"))
            rec[key] = val
        if not skip:
            rec["_row"] = ri
            mapped.append(rec)

    return {"mapped": mapped, "errors": errors, "column_mapping": mapping}

# ─────────────────────────── SQLite ファイル解析 ───────────────────────────

# テーブル名のエイリアス（POSによって違うため複数想定）
TABLE_NAME_ALIASES = {
    "casts": ["cast", "casts", "staff", "staffs", "employee", "employees",
              "girl", "girls", "hostess", "hostesses", "person", "persons",
              "member", "members", "performer", "キャスト", "スタッフ", "従業員"],
    "tables": ["table", "tables", "seat", "seats", "room", "rooms",
               "卓", "席", "テーブル", "フロア"],
    "items": ["item", "items", "product", "products", "menu", "menus",
              "menuitem", "menu_item", "menu_items", "merchandise",
              "drink", "drinks", "food", "foods", "メニュー", "商品", "ドリンク"],
    "customers": ["customer", "customers", "client", "clients",
                  "guest", "guests", "user", "users", "顧客", "お客様"],
}

def scan_sqlite_file(file_bytes: bytes) -> Dict[str, Any]:
    """SQLiteファイルをスキャンし、データ種別ごとに最適なテーブルを推定"""
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        # 読み取り専用で開く（安全のため）
        conn = sqlite3.connect(f'file:{tmp_path}?mode=ro', uri=True)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # 全テーブル取得
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
        tables = [row[0] for row in cursor.fetchall()]

        result = {
            "tables_found": [],
            "matches": {},
            "errors": [],
        }

        # 各テーブルの情報を収集
        for t in tables:
            try:
                cursor.execute(f'PRAGMA table_info("{t}")')
                columns = [row[1] for row in cursor.fetchall()]
                cursor.execute(f'SELECT COUNT(*) FROM "{t}"')
                row_count = cursor.fetchone()[0]
                result["tables_found"].append({
                    "name": t,
                    "columns": columns,
                    "row_count": row_count,
                })
            except Exception as e:
                result["errors"].append(f"テーブル {t} の解析失敗: {e}")

        # データ種別ごとに最適なテーブルを推定
        for dt_key, dt_spec in DATA_TYPES.items():
            best_table = None
            best_score = 0
            best_mapping = {}

            for t_info in result["tables_found"]:
                table_name = t_info["name"]
                columns = t_info["columns"]

                # 列マッチングでスコア計算
                mapping = {}
                score = 0
                required_matched = 0
                required_total = 0

                for field in dt_spec["fields"]:
                    if field.get("required"):
                        required_total += 1
                    idx = find_field(columns, field["aliases"])
                    if idx is not None:
                        mapping[field["key"]] = columns[idx]
                        score += 10 if field.get("required") else 5
                        if field.get("required"):
                            required_matched += 1

                # 必須項目が全部マッチしていなければ却下
                if required_total > 0 and required_matched < required_total:
                    continue

                # テーブル名ボーナス
                tn_lower = table_name.lower()
                for alias in TABLE_NAME_ALIASES.get(dt_key, []):
                    al = alias.lower()
                    if tn_lower == al:
                        score += 50  # 完全一致
                        break
                    elif al in tn_lower or tn_lower in al:
                        score += 20  # 部分一致
                        break

                if score > best_score:
                    best_score = score
                    best_table = table_name
                    best_mapping = mapping

            if best_table:
                # 該当テーブルの行数取得
                row_count = next((t["row_count"] for t in result["tables_found"] if t["name"] == best_table), 0)
                # プレビューデータも取得（最初の5件）
                preview_rows = []
                try:
                    cursor.execute(f'SELECT * FROM "{best_table}" LIMIT 5')
                    cols = [d[0] for d in cursor.description]
                    for row in cursor.fetchall():
                        preview_rows.append(dict(zip(cols, row)))
                except Exception:
                    pass
                result["matches"][dt_key] = {
                    "table": best_table,
                    "column_mapping": best_mapping,
                    "row_count": row_count,
                    "score": best_score,
                    "confidence": "high" if best_score >= 50 else ("medium" if best_score >= 25 else "low"),
                    "preview": preview_rows,
                }
        conn.close()
        return result
    except sqlite3.DatabaseError as e:
        return {"tables_found": [], "matches": {}, "errors": [f"SQLiteファイルとして読めません: {e}"]}
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

def extract_sqlite_rows(file_bytes: bytes, data_type: str, table_name: str, column_mapping: Dict[str, str]) -> List[Dict]:
    """SQLiteから指定テーブル・マッピングで行を抽出 → 内部形式に変換"""
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        conn = sqlite3.connect(f'file:{tmp_path}?mode=ro', uri=True)
        cursor = conn.cursor()
        cursor.execute(f'SELECT * FROM "{table_name}"')
        cols = [d[0] for d in cursor.description]
        rows = cursor.fetchall()
        conn.close()

        # rowsを辞書のリストに変換
        dict_rows = [dict(zip(cols, r)) for r in rows]

        # auto_map形式で正規化
        spec = DATA_TYPES[data_type]
        result = []
        for ri, row in enumerate(dict_rows, start=2):
            rec = {}
            skip = False
            for f in spec["fields"]:
                key = f["key"]
                src_col = column_mapping.get(key)
                raw_val = row.get(src_col, '') if src_col else ''
                ftype = f.get("type", "str")
                default = f.get("default", "")
                val = cast_value(raw_val, ftype, default)
                if f.get("required") and (val is None or val == ""):
                    if raw_val is None or str(raw_val).strip() == "":
                        skip = True
                        break
                if key == "category" and isinstance(val, str):
                    v = val.strip().lower()
                    val = CATEGORY_MAP.get(v, CATEGORY_MAP.get(val.strip(), "drink"))
                rec[key] = val
            if not skip:
                rec["_row"] = ri
                result.append(rec)
        return result
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

# ─────────────────────────── テンプレート生成 ───────────────────────────

def make_template_csv(data_type: str) -> str:
    """データ種別ごとのCSVテンプレートを生成"""
    spec = DATA_TYPES[data_type]
    headers = [f["label"] for f in spec["fields"]]

    # サンプル行
    samples = {
        "casts": [
            ["ゆかり", "エース", "1500", "30", "3000", "1000", "5000", "0"],
            ["なな", "レギュラー", "1300", "25", "2000", "800", "4000", "0"],
            ["みお", "新人", "1200", "20", "1500", "500", "3000", "0"],
        ],
        "tables": [["T1"], ["T2"], ["T3"], ["T4"], ["T5"]],
        "items": [
            ["1時間セット", "set", "4000", "99"],
            ["ハイボール", "drink", "800", "99"],
            ["ビール", "drink", "800", "99"],
            ["シャンパン", "bottle", "25000", "10"],
            ["おつまみ盛合せ", "food", "1500", "30"],
        ],
        "customers": [
            ["田中様", "090-1234-5678", "ウイスキー好き、誕生日5/10"],
            ["佐藤様", "080-9876-5432", "シャンパン派"],
        ],
    }

    output = io.StringIO()
    output.write('﻿')  # UTF-8 BOM（Excelで開きやすく）
    writer = csv.writer(output)
    writer.writerow(headers)
    for s in samples.get(data_type, []):
        # フィールド数を合わせる
        row = list(s) + [""] * (len(headers) - len(s))
        writer.writerow(row[:len(headers)])
    return output.getvalue()

# ─────────────────────────── API エンドポイント ───────────────────────────

@router.get("/api/import/types")
def list_types(x_role: Optional[str] = Header(None, alias="X-Role")):
    """データ種別一覧"""
    require_role(x_role, ADMIN_ROLES)
    return [
        {"key": k, "label": v["label"], "icon": v["icon"],
         "fields": [{"key": f["key"], "label": f["label"], "required": f.get("required", False)} for f in v["fields"]]}
        for k, v in DATA_TYPES.items()
    ]

@router.get("/api/import/template/{data_type}")
def download_template(data_type: str,
                       x_role: Optional[str] = Header(None, alias="X-Role")):
    """CSVテンプレートをダウンロード"""
    require_role(x_role, ADMIN_ROLES)
    if data_type not in DATA_TYPES:
        raise HTTPException(404, "不明なデータ種別")
    content = make_template_csv(data_type)
    label = DATA_TYPES[data_type]["label"]
    return StreamingResponse(
        io.BytesIO(content.encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="NEXUSCabaret_{data_type}_template.csv"'}
    )

@router.post("/api/import/preview")
async def preview_import(data_type: str = Form(...),
                          file: UploadFile = File(...),
                          x_role: Optional[str] = Header(None, alias="X-Role")):
    """ファイルをパース→プレビューデータを返す（DB変更なし）"""
    require_role(x_role, ADMIN_ROLES)
    if data_type not in DATA_TYPES:
        raise HTTPException(404, "不明なデータ種別")
    raw = await file.read()
    try:
        rows = parse_file(raw, file.filename or "data.csv")
    except Exception as e:
        raise HTTPException(400, f"ファイル読み込みエラー: {str(e)}")
    result = auto_map(rows, data_type)
    return {
        "data_type": data_type,
        "total_rows": len(rows),
        "valid_rows": len(result["mapped"]),
        "preview": result["mapped"][:20],  # 最初の20件をプレビュー
        "column_mapping": result["column_mapping"],
        "errors": result["errors"],
        "filename": file.filename,
    }

@router.post("/api/import/execute")
async def execute_import(data_type: str = Form(...),
                          file: UploadFile = File(...),
                          store_id: int = Form(1),
                          skip_duplicates: bool = Form(True),
                          x_role: Optional[str] = Header(None, alias="X-Role")):
    """ファイルから実際にDBへインポート"""
    require_role(x_role, ADMIN_ROLES)
    if data_type not in DATA_TYPES:
        raise HTTPException(404, "不明なデータ種別")
    raw = await file.read()
    try:
        rows = parse_file(raw, file.filename or "data.csv")
    except Exception as e:
        raise HTTPException(400, f"ファイル読み込みエラー: {str(e)}")
    result = auto_map(rows, data_type)
    if result["errors"]:
        raise HTTPException(400, "; ".join(result["errors"]))

    # DB保存
    from pos import Cast, Table, Item, Customer
    from cast_salary import CastSalaryConfig

    db = SessionLocal()
    inserted = 0
    updated = 0
    skipped = 0
    try:
        for rec in result["mapped"]:
            rec.pop("_row", None)
            if data_type == "casts":
                existing = db.query(Cast).filter_by(store_id=store_id, name=rec["name"]).first()
                if existing:
                    if skip_duplicates:
                        skipped += 1
                        continue
                    cast = existing
                    updated += 1
                else:
                    cast = Cast(store_id=store_id, name=rec["name"],
                                rank=rec.get("rank", ""), is_active=True)
                    db.add(cast)
                    db.flush()
                    inserted += 1
                # 給与設定（CastSalaryConfig）も同時に作成・更新
                cfg = db.query(CastSalaryConfig).filter_by(cast_id=cast.id).first()
                if not cfg:
                    cfg = CastSalaryConfig(cast_id=cast.id, store_id=store_id)
                    db.add(cfg)
                for k in ["hourly_rate", "drink_back_rate", "nom_fee_hon",
                          "nom_fee_jyonai", "nom_fee_dohan", "floor_rate"]:
                    if k in rec and rec[k] is not None:
                        try:
                            setattr(cfg, k, float(rec[k]))
                        except (ValueError, TypeError):
                            pass
            elif data_type == "tables":
                existing = db.query(Table).filter_by(store_id=store_id, name=rec["name"]).first()
                if existing:
                    if skip_duplicates:
                        skipped += 1
                        continue
                else:
                    db.add(Table(store_id=store_id, name=rec["name"]))
                    inserted += 1
            elif data_type == "items":
                existing = db.query(Item).filter_by(store_id=store_id, name=rec["name"]).first()
                if existing:
                    if skip_duplicates:
                        skipped += 1
                        continue
                    item = existing
                    updated += 1
                else:
                    item = Item(store_id=store_id, name=rec["name"])
                    db.add(item)
                    inserted += 1
                item.category = rec.get("category", "drink")
                item.price = float(rec.get("price", 0) or 0)
                item.stock = int(rec.get("stock", 0) or 0)
                if item.category == "bottle":
                    item.keepable = True
            elif data_type == "customers":
                existing = db.query(Customer).filter_by(store_id=store_id, nickname=rec["nickname"]).first()
                if existing:
                    if skip_duplicates:
                        skipped += 1
                        continue
                    cust = existing
                    updated += 1
                else:
                    cust = Customer(store_id=store_id, nickname=rec["nickname"])
                    db.add(cust)
                    inserted += 1
                cust.phone = rec.get("phone", "") or ""
                cust.memo = rec.get("memo", "") or ""
        db.commit()
        return {"ok": True, "inserted": inserted, "updated": updated, "skipped": skipped,
                "total": inserted + updated + skipped}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"インポート中エラー: {str(e)}")
    finally:
        db.close()

# ─────────────────────────── SQLite API ───────────────────────────

@router.post("/api/import/sqlite/scan")
async def sqlite_scan(file: UploadFile = File(...),
                       x_role: Optional[str] = Header(None, alias="X-Role")):
    """アップロードされたSQLiteファイルをスキャン → 全データ種別の検出結果を返す"""
    require_role(x_role, ADMIN_ROLES)
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "ファイルが空です")
    result = scan_sqlite_file(raw)
    if not result["tables_found"]:
        raise HTTPException(400, "SQLiteファイルにテーブルが見つかりません。"
                                  "（ファイル形式が間違っている可能性があります）")
    return {
        "filename": file.filename,
        "tables_count": len(result["tables_found"]),
        "tables_list": [{"name": t["name"], "columns": t["columns"], "row_count": t["row_count"]} for t in result["tables_found"]],
        "matches": result["matches"],
        "errors": result["errors"],
    }

@router.post("/api/import/sqlite/execute")
async def sqlite_execute(
    file: UploadFile = File(...),
    selected_types: str = Form(...),  # カンマ区切り: "casts,items,customers"
    store_id: int = Form(1),
    skip_duplicates: bool = Form(True),
    x_role: Optional[str] = Header(None, alias="X-Role"),
):
    """SQLiteファイルから複数データ種別を一括インポート"""
    require_role(x_role, ADMIN_ROLES)
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "ファイルが空です")
    scan = scan_sqlite_file(raw)
    types_to_import = [t.strip() for t in selected_types.split(",") if t.strip() in DATA_TYPES]
    if not types_to_import:
        raise HTTPException(400, "取り込み対象が選択されていません")

    # DB保存ループ
    from pos import Cast, Table, Item, Customer
    from cast_salary import CastSalaryConfig

    db = SessionLocal()
    summary = {}
    try:
        for dt_key in types_to_import:
            match = scan["matches"].get(dt_key)
            if not match:
                summary[dt_key] = {"inserted": 0, "updated": 0, "skipped": 0, "skipped_reason": "対象テーブル未検出"}
                continue

            rows = extract_sqlite_rows(raw, dt_key, match["table"], match["column_mapping"])
            inserted = updated = skipped = 0

            for rec in rows:
                rec.pop("_row", None)
                if dt_key == "casts":
                    existing = db.query(Cast).filter_by(store_id=store_id, name=rec["name"]).first()
                    if existing:
                        if skip_duplicates:
                            skipped += 1; continue
                        cast = existing; updated += 1
                    else:
                        cast = Cast(store_id=store_id, name=rec["name"],
                                    rank=rec.get("rank", ""), is_active=True)
                        db.add(cast); db.flush(); inserted += 1
                    cfg = db.query(CastSalaryConfig).filter_by(cast_id=cast.id).first()
                    if not cfg:
                        cfg = CastSalaryConfig(cast_id=cast.id, store_id=store_id)
                        db.add(cfg)
                    for k in ["hourly_rate","drink_back_rate","nom_fee_hon","nom_fee_jyonai","nom_fee_dohan","floor_rate"]:
                        if k in rec and rec[k] is not None:
                            try: setattr(cfg, k, float(rec[k]))
                            except (ValueError, TypeError): pass
                elif dt_key == "tables":
                    existing = db.query(Table).filter_by(store_id=store_id, name=rec["name"]).first()
                    if existing:
                        if skip_duplicates: skipped += 1; continue
                    else:
                        db.add(Table(store_id=store_id, name=rec["name"])); inserted += 1
                elif dt_key == "items":
                    existing = db.query(Item).filter_by(store_id=store_id, name=rec["name"]).first()
                    if existing:
                        if skip_duplicates: skipped += 1; continue
                        item = existing; updated += 1
                    else:
                        item = Item(store_id=store_id, name=rec["name"]); db.add(item); inserted += 1
                    item.category = rec.get("category", "drink")
                    item.price = float(rec.get("price", 0) or 0)
                    item.stock = int(rec.get("stock", 0) or 0)
                    if item.category == "bottle":
                        item.keepable = True
                elif dt_key == "customers":
                    existing = db.query(Customer).filter_by(store_id=store_id, nickname=rec["nickname"]).first()
                    if existing:
                        if skip_duplicates: skipped += 1; continue
                        cust = existing; updated += 1
                    else:
                        cust = Customer(store_id=store_id, nickname=rec["nickname"]); db.add(cust); inserted += 1
                    cust.phone = rec.get("phone", "") or ""
                    cust.memo = rec.get("memo", "") or ""

            summary[dt_key] = {
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped,
                "source_table": match["table"],
            }
        db.commit()
        return {"ok": True, "summary": summary}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"インポート中エラー: {str(e)}")
    finally:
        db.close()

# ─────────────────────────── UI ───────────────────────────

@router.get("/ui/import", response_class=HTMLResponse)
def ui_import():
    """データ移行ウィザード画面"""
    return HTMLResponse(r"""<!doctype html>
<html lang="ja"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>データ移行ウィザード - NEXUS Cabaret</title>
<style>
:root{--bg:#fafafa;--card:#ffffff;--card2:#ffffff;--line:#eaeaef;--border:#eaeaef;--text:#0a0a0f;--ink:#0a0a0f;--muted:#8a8a95;--body:#4a4a55;--accent:#d64583;--accent-soft:#fdf0f7;--accent-dark:#b03468;--gold:#c9a96e;--gold-soft:#faf3e3;--warn:#f59e0b;--amber:#f59e0b;--err:#ef4444;--red:#ef4444;--ok:#22c55e;--green:#22c55e;--blue:#3b82f6;--purple:#a855f7;}
*{box-sizing:border-box;font-family:-apple-system,system-ui,"Noto Sans JP",sans-serif;margin:0;padding:0}
body{background:var(--bg);color:var(--text);line-height:1.7}
header{position:sticky;top:0;z-index:40;display:flex;align-items:center;gap:12px;padding:14px 18px;border-bottom:1px solid var(--line);background:rgba(11,18,32,.95);backdrop-filter:blur(8px)}
header h1{font-size:17px;margin:0}
.nav{margin-left:auto;display:flex;gap:8px}
.nav a{color:var(--muted);text-decoration:none;font-size:13px;padding:6px 10px;border-radius:8px}
.nav a:hover{color:var(--text);background:#1a2438}
.container{max-width:880px;margin:0 auto;padding:24px 16px;display:flex;flex-direction:column;gap:18px}
.card{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:22px}
.card h2{font-size:15px;margin-bottom:14px;border-bottom:1px solid var(--line);padding-bottom:10px}

/* ステップインジケーター */
.steps{display:flex;justify-content:space-between;margin-bottom:24px;position:relative}
.steps::before{content:'';position:absolute;top:18px;left:10%;right:10%;height:2px;background:var(--line);z-index:0}
.step{flex:1;text-align:center;position:relative;z-index:1}
.step-num{display:inline-flex;align-items:center;justify-content:center;width:36px;height:36px;border-radius:50%;background:var(--line);color:var(--muted);font-weight:700;font-size:14px;border:2px solid var(--bg);transition:.3s}
.step.on .step-num{background:var(--accent);color:#001018}
.step.done .step-num{background:var(--green);color:#fff}
.step-label{font-size:11px;color:var(--muted);margin-top:6px}
.step.on .step-label,.step.done .step-label{color:var(--text)}

/* データ種別選択カード */
.type-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:12px}
.type-card{background:#0a1423;border:2px solid var(--line);border-radius:12px;padding:18px;cursor:pointer;transition:.2s;text-align:left}
.type-card:hover{border-color:var(--accent);transform:translateY(-2px)}
.type-card.on{border-color:var(--accent);background:#0c2a3d}
.type-card .icon{font-size:28px;margin-bottom:8px;display:block}
.type-card .name{font-weight:700;font-size:15px;margin-bottom:4px}
.type-card .desc{font-size:11px;color:var(--muted)}

/* ボタン */
.btn{cursor:pointer;font-size:14px;padding:11px 22px;border-radius:10px;border:1px solid #334155;background:#111827;color:var(--text);font-weight:600;display:inline-flex;align-items:center;gap:8px;text-decoration:none;transition:.2s}
.btn:hover{transform:translateY(-1px)}
.btn.primary{background:linear-gradient(135deg,#0ea5e9,#0284c7);border-color:#0ea5e9;color:#fff}
.btn.success{background:linear-gradient(135deg,#22c55e,#16a34a);border-color:#22c55e;color:#fff}
.btn.ghost{background:transparent;color:var(--muted)}
.btn.ghost:hover{color:var(--text)}
.btn:disabled{opacity:.5;cursor:not-allowed;transform:none}
.btn.full{width:100%;justify-content:center}

/* ドロップゾーン */
.drop{border:2px dashed var(--line);border-radius:12px;padding:40px 20px;text-align:center;cursor:pointer;transition:.2s;background:#0a1423}
.drop:hover,.drop.over{border-color:var(--accent);background:#0c2a3d}
.drop input{display:none}
.drop .ic{font-size:42px;display:block;margin-bottom:10px}
.drop .ti{font-size:14px;font-weight:600;margin-bottom:4px}
.drop .sub{font-size:11px;color:var(--muted)}

/* プレビューテーブル */
.preview-tbl{width:100%;border-collapse:collapse;font-size:12px;background:#0a1423;border-radius:8px;overflow:hidden}
.preview-tbl th,.preview-tbl td{padding:8px 10px;border-bottom:1px solid var(--line);text-align:left}
.preview-tbl th{background:#111827;font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px}
.preview-tbl tr:last-child td{border-bottom:none}

/* メッセージ */
.alert{padding:14px 16px;border-radius:10px;font-size:14px;margin-bottom:14px;display:flex;gap:10px;align-items:flex-start}
.alert.success{background:#0f2615;border:1px solid var(--green);color:#86efac}
.alert.error{background:#1a0e12;border:1px solid var(--red);color:#fca5a5}
.alert.info{background:#0c1d2e;border:1px solid var(--accent);color:#7dd3fc}
.alert.warn{background:#1f1707;border:1px solid var(--amber);color:#fcd34d}

/* マッピング表示 */
.mapping{background:#0a1423;border:1px solid var(--line);border-radius:8px;padding:12px;margin-top:12px;font-size:12px}
.mapping-row{display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px dotted var(--line)}
.mapping-row:last-child{border:none}
.mapping-row .src{color:var(--muted)}
.mapping-row .arr{color:var(--accent)}
.mapping-row .dst{color:var(--text);font-weight:600}

.hidden{display:none!important}

/* 方法選択タブ */
.method-tabs{display:flex;gap:8px;margin-bottom:18px;background:#0a1423;border:1px solid var(--line);border-radius:12px;padding:6px}
.method-tab{flex:1;padding:14px 18px;border-radius:8px;border:none;background:transparent;color:var(--muted);font-weight:600;cursor:pointer;font-size:13px;transition:.2s;font-family:inherit}
.method-tab:hover{color:var(--text)}
.method-tab.on{background:var(--accent);color:#001018}

/* SQLite検出結果カード */
.match-card{background:#0a1423;border:1px solid var(--line);border-radius:10px;padding:14px 16px;margin-bottom:10px;display:flex;align-items:center;gap:12px;transition:.2s}
.match-card.high{border-color:var(--green)}
.match-card.medium{border-color:var(--amber)}
.match-card.low{border-color:var(--red);opacity:.65}
.match-card.none{border-color:var(--line);opacity:.45}
.match-toggle{flex-shrink:0}
.match-toggle input{width:20px;height:20px;cursor:pointer;accent-color:var(--accent)}
.match-info{flex:1;min-width:0}
.match-info .name{font-weight:700;font-size:14px;margin-bottom:2px}
.match-info .meta{font-size:11px;color:var(--muted)}
.match-info .meta strong{color:var(--text)}
.confidence-pill{display:inline-block;padding:2px 8px;border-radius:999px;font-size:10px;font-weight:700;letter-spacing:.05em}
.confidence-pill.high{background:rgba(34,197,94,.15);color:#86efac}
.confidence-pill.medium{background:rgba(245,158,11,.15);color:#fcd34d}
.confidence-pill.low{background:rgba(239,68,68,.15);color:#fca5a5}
.confidence-pill.none{background:#1c1c2e;color:var(--muted)}

@media(max-width:600px){
  .container{padding:14px 10px}
  .type-grid{grid-template-columns:1fr}
  .card{padding:16px}
  .step-label{font-size:9px}
  .step-num{width:30px;height:30px;font-size:12px}
}

/* === Premium Pink Theme Override (auto-injected) === */
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;500;700&family=Inter:wght@400;500;600;700;800&display=swap');
body{font-family:'Inter','Noto Sans JP',-apple-system,system-ui,Segoe UI,Roboto,sans-serif !important;background:#fafafa !important;color:#0a0a0f !important;-webkit-font-smoothing:antialiased}
h1,h2,h3,h4{color:#0a0a0f}
h1{font-weight:800;letter-spacing:-.01em}
a{color:#d64583}
.card,.box,section,article{background:#ffffff !important;border-color:#eaeaef !important;color:#0a0a0f}
.stat,.kpi,.tile,.metric{background:#ffffff !important;border:1px solid #eaeaef !important;color:#0a0a0f}
.stat .val,.kpi .val,.metric .val,.tile .val{color:#0a0a0f !important}
.stat .label,.kpi .label,.tile .label{color:#8a8a95 !important}
table{border-color:#eaeaef !important}
th{background:#fafafa !important;color:#4a4a55 !important;font-weight:700 !important;border-color:#eaeaef !important;letter-spacing:.02em}
td{border-color:#f3f3f6 !important;color:#0a0a0f !important;background:#ffffff}
tr:nth-child(even) td{background:#fafafa}
.badge.open{background:#fafafa !important;color:#8a8a95 !important;border:1px solid #eaeaef !important}
.badge.preliminary{background:#fff7ed !important;color:#c2410c !important;border-color:#fed7aa !important}
.badge.final{background:#f0fdf4 !important;color:#15803d !important;border-color:#86efac !important}
.btn{background:#ffffff !important;border:1px solid #eaeaef !important;color:#0a0a0f !important;transition:all .2s;font-weight:600 !important}
.btn:hover{border-color:#d64583 !important;color:#d64583 !important;background:#ffffff !important;transform:translateY(-1px);box-shadow:0 2px 8px rgba(214,69,131,.08)}
.btn.primary,.btn.solid{background:#d64583 !important;border-color:#d64583 !important;color:#ffffff !important}
.btn.primary:hover,.btn.solid:hover{background:#b03468 !important;border-color:#b03468 !important;color:#ffffff !important;opacity:1 !important}
.btn.warn{background:#f59e0b !important;border-color:#f59e0b !important;color:#ffffff !important}
.btn.danger,.btn.err{background:#ef4444 !important;border-color:#ef4444 !important;color:#ffffff !important}
input,select,textarea{background:#ffffff !important;border:1px solid #eaeaef !important;color:#0a0a0f !important;font-family:inherit}
input:focus,select:focus,textarea:focus{border-color:#d64583 !important;box-shadow:0 0 0 3px #fdf0f7 !important;outline:none}
.tab{color:#8a8a95 !important;background:transparent !important}
.tab.active{background:#fdf0f7 !important;color:#d64583 !important;border-color:#eaeaef !important;border-bottom-color:#fdf0f7 !important;font-weight:700}
.tab-body{background:#ffffff !important;border-color:#eaeaef !important}
.toast.ok{background:#f0fdf4 !important;color:#14532d !important;border:1px solid #86efac}
.toast.err{background:#fef2f2 !important;color:#7f1d1d !important;border:1px solid #fca5a5}
.history-row{border-color:#f3f3f6 !important}
header{background:rgba(255,255,255,.92) !important;backdrop-filter:blur(20px);border-color:#eaeaef !important;color:#0a0a0f !important}
header h1{background:linear-gradient(135deg,#0a0a0f,#d64583);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
nav a,.nav a{color:#4a4a55}
nav a:hover,.nav a:hover{color:#d64583}
/* よくあるダーク背景の hex 値を強制ライト化 */
[style*="background:#0f172a"],[style*="background:#0b1220"],[style*="background:#1e293b"],[style*="background:#0a1423"],[style*="background:#111827"],[style*="background:#1a1d29"]{background:#ffffff !important;color:#0a0a0f !important;border-color:#eaeaef}
[style*="color:#e5e7eb"],[style*="color:#e2e8f0"]{color:#0a0a0f !important}
[style*="color:#94a3b8"],[style*="color:#9ca3af"]{color:#8a8a95 !important}
[style*="border:1px solid #1f2937"],[style*="border:1px solid #334155"],[style*="border:1px solid #263244"]{border-color:#eaeaef !important}

</style></head><body>
<header>
  <h1>📦 データ移行ウィザード</h1>
  <div class="nav">
    <a href="/ui">← フロア</a>
    <a href="/ui/casts">キャスト</a>
    <a href="/ui/items">メニュー</a>
    <a href="/ui/customers">顧客</a>
  </div>
</header>

<div class="container">

  <!-- インポート方法選択 -->
  <div class="method-tabs">
    <button class="method-tab on" data-method="csv" onclick="switchMethod('csv')">📄 CSV / Excel ファイル</button>
    <button class="method-tab" data-method="sqlite" onclick="switchMethod('sqlite')">💾 データベースファイル (.db)</button>
  </div>

  <!-- 通知（共通） -->
  <div id="alertBox"></div>

  <!-- ═════════ CSV / Excel フロー ═════════ -->
  <div id="csvFlow">

  <!-- ステップ -->
  <div class="steps">
    <div class="step on" id="step1"><div class="step-num">1</div><div class="step-label">種別選択</div></div>
    <div class="step" id="step2"><div class="step-num">2</div><div class="step-label">ファイル選択</div></div>
    <div class="step" id="step3"><div class="step-num">3</div><div class="step-label">プレビュー</div></div>
    <div class="step" id="step4"><div class="step-num">4</div><div class="step-label">完了</div></div>
  </div>

  <!-- ─── Step 1: 種別選択 ─── -->
  <div class="card" id="card1">
    <h2>① インポートするデータを選んでください</h2>
    <div class="type-grid">
      <button class="type-card" data-type="casts">
        <span class="icon">👤</span>
        <div class="name">キャスト</div>
        <div class="desc">氏名・時給・各種バック</div>
      </button>
      <button class="type-card" data-type="tables">
        <span class="icon">🪑</span>
        <div class="name">テーブル</div>
        <div class="desc">卓名・席名</div>
      </button>
      <button class="type-card" data-type="items">
        <span class="icon">🍸</span>
        <div class="name">メニュー（商品）</div>
        <div class="desc">商品名・カテゴリ・価格</div>
      </button>
      <button class="type-card" data-type="customers">
        <span class="icon">🤝</span>
        <div class="name">顧客</div>
        <div class="desc">ニックネーム・連絡先・メモ</div>
      </button>
    </div>
  </div>

  <!-- ─── Step 2: ファイルアップロード ─── -->
  <div class="card hidden" id="card2">
    <h2>② <span id="selectedTypeLabel"></span>のデータをアップロード</h2>
    <div class="alert info">
      💡 既存POSのエクスポートCSV、Excel（.xlsx）、テンプレートのいずれでもOK。<br>
      列名は自動でマッチングします（日本語/英語対応）。
    </div>
    <div style="margin-bottom:14px;display:flex;gap:8px;flex-wrap:wrap">
      <button class="btn ghost" id="btnDownloadTemplate">
        📥 テンプレートをダウンロード（推奨）
      </button>
      <button class="btn ghost" id="btnBack1">← データ種別を選び直す</button>
    </div>
    <label class="drop" id="drop">
      <input type="file" id="fileInput" accept=".csv,.xlsx,.xls,.tsv,.txt">
      <span class="ic">📂</span>
      <div class="ti">クリックしてファイルを選択 または ドラッグ&ドロップ</div>
      <div class="sub">CSV / Excel(.xlsx) / TSV 対応</div>
    </label>
    <div id="fileName" style="margin-top:12px;text-align:center;color:var(--muted);font-size:13px"></div>
  </div>

  <!-- ─── Step 3: プレビュー ─── -->
  <div class="card hidden" id="card3">
    <h2>③ 取り込み内容を確認</h2>
    <div id="previewSummary"></div>
    <div id="mappingArea" class="mapping hidden"></div>
    <div style="overflow-x:auto;margin-top:14px;max-height:400px;overflow-y:auto">
      <table class="preview-tbl" id="previewTbl"></table>
    </div>
    <div style="margin-top:18px;display:flex;gap:8px;flex-wrap:wrap;justify-content:space-between">
      <button class="btn ghost" id="btnBack2">← ファイルを選び直す</button>
      <button class="btn success" id="btnExecute">✅ この内容でインポート実行</button>
    </div>
  </div>

  <!-- ─── Step 4: 完了 ─── -->
  <div class="card hidden" id="card4">
    <h2>④ インポート完了</h2>
    <div id="completionArea"></div>
    <div style="margin-top:20px;display:flex;gap:8px;flex-wrap:wrap">
      <button class="btn primary" onclick="resetAll()">🔄 別のデータをインポート</button>
      <a class="btn ghost" href="/ui">フロアに戻る</a>
    </div>
  </div>

  </div><!-- /csvFlow -->

  <!-- ═════════ SQLite フロー ═════════ -->
  <div id="sqliteFlow" class="hidden">

    <!-- Step 1: ファイル選択 -->
    <div class="card" id="sqCard1">
      <h2>💾 SQLiteデータベースファイルから一括移行</h2>
      <div class="alert info">
        💡 既存POSの <strong>.db / .sqlite / .sqlite3</strong> ファイルをアップロードしてください。<br>
        テーブル構造を自動解析して、キャスト・テーブル・メニュー・顧客を**一発で**取り込みます。
      </div>
      <label class="drop" id="sqDrop">
        <input type="file" id="sqFileInput" accept=".db,.sqlite,.sqlite3">
        <span class="ic">💾</span>
        <div class="ti">クリックしてSQLiteファイルを選択 または ドラッグ&ドロップ</div>
        <div class="sub">.db / .sqlite / .sqlite3 対応</div>
      </label>
      <div id="sqFileName" style="margin-top:12px;text-align:center;color:var(--muted);font-size:13px"></div>
    </div>

    <!-- Step 2: 自動検出結果 -->
    <div class="card hidden" id="sqCard2">
      <h2>🔍 自動検出結果</h2>
      <div id="sqScanSummary"></div>
      <div id="sqMatchesArea" style="margin-top:14px"></div>
      <div style="margin-top:18px;display:flex;gap:8px;flex-wrap:wrap;justify-content:space-between">
        <button class="btn ghost" onclick="resetSqlite()">← ファイルを選び直す</button>
        <button class="btn success" id="sqExecute">✅ チェックを入れたデータを一括インポート</button>
      </div>
    </div>

    <!-- Step 3: 完了 -->
    <div class="card hidden" id="sqCard3">
      <h2>🎉 インポート完了</h2>
      <div id="sqCompletion"></div>
      <div style="margin-top:20px;display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn primary" onclick="resetSqlite()">🔄 別のファイルをインポート</button>
        <a class="btn ghost" href="/ui">フロアに戻る</a>
      </div>
    </div>

  </div><!-- /sqliteFlow -->

</div>

<script>
const $ = id => document.getElementById(id);
let selectedType = null;
let selectedFile = null;
let previewData = null;
const TYPE_LABELS = {casts:'キャスト', tables:'テーブル', items:'メニュー', customers:'顧客'};

function showAlert(type, msg, ms=6000){
  const a = $('alertBox');
  a.innerHTML = `<div class="alert ${type}">${msg}</div>`;
  if(ms) setTimeout(()=>{ a.innerHTML = ''; }, ms);
}

function goStep(n){
  for(let i=1;i<=4;i++){
    $(`card${i}`).classList.toggle('hidden', i !== n);
    const s = $(`step${i}`);
    s.classList.remove('on', 'done');
    if(i < n) s.classList.add('done');
    if(i === n) s.classList.add('on');
  }
  window.scrollTo({top:0, behavior:'smooth'});
}

async function api(path, opt={}){
  const tk = sessionStorage.getItem('pos_token')||'';
  const o = {method:'GET', headers:{'X-Role':'owner','X-Token':tk}, ...opt};
  if(!opt.body || !(opt.body instanceof FormData)){
    o.headers['Content-Type'] = 'application/json';
    if(o.body && typeof o.body !== 'string') o.body = JSON.stringify(o.body);
  }
  const r = await fetch(path, o);
  if(r.status === 401){ sessionStorage.clear(); window.location.href='/'; return; }
  if(!r.ok) throw new Error(await r.text());
  const ct = r.headers.get('content-type')||'';
  return ct.includes('json') ? r.json() : r.text();
}

// ─── Step 1: 種別選択 ───
document.querySelectorAll('.type-card').forEach(c => {
  c.addEventListener('click', () => {
    document.querySelectorAll('.type-card').forEach(x => x.classList.remove('on'));
    c.classList.add('on');
    selectedType = c.dataset.type;
    $('selectedTypeLabel').textContent = TYPE_LABELS[selectedType] || selectedType;
    setTimeout(() => goStep(2), 200);
  });
});

// ─── Step 2: テンプレートDL & ファイル選択 ───
$('btnDownloadTemplate').addEventListener('click', () => {
  const tk = sessionStorage.getItem('pos_token')||'';
  const url = `/api/import/template/${selectedType}`;
  // headersをセットできないのでcookieのみで認証
  fetch(url, {headers:{'X-Role':'owner','X-Token':tk}}).then(r => r.blob()).then(b => {
    const a = document.createElement('a');
    a.href = URL.createObjectURL(b);
    a.download = `NEXUSCabaret_${selectedType}_template.csv`;
    a.click();
  }).catch(e => showAlert('error','テンプレートDL失敗: ' + e.message));
});

$('btnBack1').addEventListener('click', () => goStep(1));
$('btnBack2').addEventListener('click', () => goStep(2));

const drop = $('drop');
const fileInput = $('fileInput');
['dragenter','dragover'].forEach(e => drop.addEventListener(e, ev => { ev.preventDefault(); drop.classList.add('over'); }));
['dragleave','drop'].forEach(e => drop.addEventListener(e, ev => { ev.preventDefault(); drop.classList.remove('over'); }));
drop.addEventListener('drop', ev => {
  if(ev.dataTransfer.files.length) handleFile(ev.dataTransfer.files[0]);
});
fileInput.addEventListener('change', () => {
  if(fileInput.files.length) handleFile(fileInput.files[0]);
});

async function handleFile(f){
  selectedFile = f;
  $('fileName').textContent = `📄 ${f.name} （${(f.size/1024).toFixed(1)} KB）`;
  // 自動でプレビュー
  const fd = new FormData();
  fd.append('data_type', selectedType);
  fd.append('file', f);
  showAlert('info', '⏳ ファイル解析中...', 0);
  try {
    const r = await fetch('/api/import/preview', {
      method:'POST',
      headers:{'X-Role':'owner','X-Token':sessionStorage.getItem('pos_token')||''},
      body: fd
    });
    if(!r.ok) throw new Error(await r.text());
    previewData = await r.json();
    $('alertBox').innerHTML = '';
    renderPreview();
    goStep(3);
  } catch(e){
    showAlert('error', '解析エラー: ' + e.message);
  }
}

function renderPreview(){
  const p = previewData;
  let summary = '';
  if(p.errors && p.errors.length){
    summary += `<div class="alert error">❌ エラー:<br>` + p.errors.join('<br>') + '</div>';
    $('btnExecute').disabled = true;
  } else {
    summary += `<div class="alert success">✅ <strong>${p.valid_rows}</strong>件のデータを取り込めます（全${p.total_rows}行中）</div>`;
    $('btnExecute').disabled = false;
  }
  $('previewSummary').innerHTML = summary;

  // 列マッピング表示
  const ma = $('mappingArea');
  if(p.column_mapping && Object.keys(p.column_mapping).length){
    ma.classList.remove('hidden');
    ma.innerHTML = '<div style="font-weight:700;margin-bottom:8px">📋 列マッピング（自動判定）</div>' +
      Object.entries(p.column_mapping).map(([k,v]) =>
        `<div class="mapping-row"><span class="src">「${v}」</span><span class="arr">→</span><span class="dst">${k}</span></div>`
      ).join('');
  } else {
    ma.classList.add('hidden');
  }

  // プレビューテーブル
  const tbl = $('previewTbl');
  if(!p.preview.length){
    tbl.innerHTML = '<tr><td>表示できるデータがありません</td></tr>';
    return;
  }
  const keys = Object.keys(p.preview[0]).filter(k => k !== '_row');
  tbl.innerHTML = '<thead><tr><th>#</th>' + keys.map(k => `<th>${k}</th>`).join('') + '</tr></thead>' +
    '<tbody>' + p.preview.map(r =>
      '<tr><td style="color:var(--muted)">' + (r._row||'') + '</td>' +
      keys.map(k => `<td>${escapeHtml(String(r[k] ?? ''))}</td>`).join('') + '</tr>'
    ).join('') + '</tbody>';
}

function escapeHtml(s){
  return s.replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}

// ─── Step 3: インポート実行 ───
$('btnExecute').addEventListener('click', async () => {
  if(!selectedFile) return;
  if(!confirm(`${previewData.valid_rows}件のデータをインポートします。よろしいですか？\n\n（既に同じ名前のデータがある場合はスキップされます）`)) return;
  const fd = new FormData();
  fd.append('data_type', selectedType);
  fd.append('file', selectedFile);
  fd.append('store_id', '1');
  fd.append('skip_duplicates', 'true');
  $('btnExecute').disabled = true;
  $('btnExecute').textContent = '⏳ インポート中...';
  try {
    const r = await fetch('/api/import/execute', {
      method:'POST',
      headers:{'X-Role':'owner','X-Token':sessionStorage.getItem('pos_token')||''},
      body: fd
    });
    if(!r.ok) throw new Error(await r.text());
    const result = await r.json();
    $('completionArea').innerHTML = `
      <div class="alert success">
        🎉 インポート完了！
      </div>
      <div style="background:#0a1423;border:1px solid var(--line);border-radius:10px;padding:18px;margin-top:14px">
        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;text-align:center">
          <div>
            <div style="font-size:32px;font-weight:700;color:var(--green)">${result.inserted}</div>
            <div style="font-size:12px;color:var(--muted)">新規追加</div>
          </div>
          <div>
            <div style="font-size:32px;font-weight:700;color:var(--accent)">${result.updated}</div>
            <div style="font-size:12px;color:var(--muted)">更新</div>
          </div>
          <div>
            <div style="font-size:32px;font-weight:700;color:var(--muted)">${result.skipped}</div>
            <div style="font-size:12px;color:var(--muted)">スキップ（既存）</div>
          </div>
        </div>
      </div>
    `;
    goStep(4);
  } catch(e){
    showAlert('error', 'インポート失敗: ' + e.message);
    $('btnExecute').disabled = false;
    $('btnExecute').textContent = '✅ この内容でインポート実行';
  }
});

function resetAll(){
  selectedType = null; selectedFile = null; previewData = null;
  document.querySelectorAll('.type-card').forEach(x => x.classList.remove('on'));
  $('fileInput').value = '';
  $('fileName').textContent = '';
  $('btnExecute').disabled = false;
  $('btnExecute').textContent = '✅ この内容でインポート実行';
  goStep(1);
}

// ═════════ 方法切替 ═════════
function switchMethod(m){
  document.querySelectorAll('.method-tab').forEach(t => t.classList.toggle('on', t.dataset.method === m));
  $('csvFlow').classList.toggle('hidden', m !== 'csv');
  $('sqliteFlow').classList.toggle('hidden', m !== 'sqlite');
  $('alertBox').innerHTML = '';
}

// ═════════ SQLite フロー ═════════
let sqFile = null;
let sqScanData = null;
const TYPE_ICONS = {casts:'👤', tables:'🪑', items:'🍸', customers:'🤝'};
const CONF_LABELS = {high:'高精度', medium:'中精度', low:'低精度', none:'未検出'};

const sqDrop = $('sqDrop');
const sqFileInput = $('sqFileInput');
['dragenter','dragover'].forEach(e => sqDrop.addEventListener(e, ev => { ev.preventDefault(); sqDrop.classList.add('over'); }));
['dragleave','drop'].forEach(e => sqDrop.addEventListener(e, ev => { ev.preventDefault(); sqDrop.classList.remove('over'); }));
sqDrop.addEventListener('drop', ev => {
  if(ev.dataTransfer.files.length) handleSqFile(ev.dataTransfer.files[0]);
});
sqFileInput.addEventListener('change', () => {
  if(sqFileInput.files.length) handleSqFile(sqFileInput.files[0]);
});

async function handleSqFile(f){
  sqFile = f;
  $('sqFileName').textContent = `💾 ${f.name} （${(f.size/1024).toFixed(1)} KB）`;
  const fd = new FormData();
  fd.append('file', f);
  showAlert('info', '⏳ データベースを解析中... テーブル構造を検出しています', 0);
  try{
    const r = await fetch('/api/import/sqlite/scan', {
      method:'POST',
      headers:{'X-Role':'owner','X-Token':sessionStorage.getItem('pos_token')||''},
      body: fd
    });
    if(!r.ok) throw new Error(await r.text());
    sqScanData = await r.json();
    $('alertBox').innerHTML = '';
    renderSqScan();
    $('sqCard2').classList.remove('hidden');
    $('sqCard2').scrollIntoView({behavior:'smooth'});
  }catch(e){
    showAlert('error', '解析エラー: ' + e.message);
  }
}

function renderSqScan(){
  const d = sqScanData;
  $('sqScanSummary').innerHTML = `
    <div class="alert info">
      📊 <strong>${d.tables_count}</strong>個のテーブルが見つかりました（${d.filename}）。<br>
      以下から取り込みたいデータにチェックを入れてください。
    </div>
  `;

  let html = '';
  const allTypes = ['casts', 'tables', 'items', 'customers'];
  for(const dt of allTypes){
    const m = d.matches[dt];
    const label = TYPE_LABELS[dt];
    const icon = TYPE_ICONS[dt];
    if(m){
      const conf = m.confidence;
      const cols = Object.entries(m.column_mapping)
        .map(([k,v]) => `<span style="color:var(--muted)">${v}</span>→<span style="color:var(--accent)">${k}</span>`)
        .join(' ／ ');
      html += `
        <div class="match-card ${conf}">
          <label class="match-toggle">
            <input type="checkbox" checked data-type="${dt}">
          </label>
          <div class="match-info">
            <div class="name">${icon} ${label}　<span class="confidence-pill ${conf}">${CONF_LABELS[conf]}</span></div>
            <div class="meta">
              📋 テーブル: <strong>${m.table}</strong>　／
              件数: <strong>${m.row_count}</strong>件
            </div>
            <div class="meta" style="margin-top:4px">列マッピング: ${cols}</div>
          </div>
        </div>
      `;
    } else {
      html += `
        <div class="match-card none">
          <label class="match-toggle">
            <input type="checkbox" disabled data-type="${dt}">
          </label>
          <div class="match-info">
            <div class="name">${icon} ${label}　<span class="confidence-pill none">未検出</span></div>
            <div class="meta">該当するテーブルが見つかりませんでした</div>
          </div>
        </div>
      `;
    }
  }
  $('sqMatchesArea').innerHTML = html;
}

$('sqExecute').addEventListener('click', async () => {
  const checked = Array.from(document.querySelectorAll('#sqMatchesArea input[type="checkbox"]:checked'))
    .map(c => c.dataset.type);
  if(!checked.length){ alert('取り込むデータを選択してください'); return; }
  if(!confirm(`選択したデータ（${checked.length}種類）を一括インポートします。よろしいですか？\n\n（既に同じ名前のデータがある場合はスキップされます）`)) return;

  const fd = new FormData();
  fd.append('file', sqFile);
  fd.append('selected_types', checked.join(','));
  fd.append('store_id', '1');
  fd.append('skip_duplicates', 'true');
  $('sqExecute').disabled = true;
  $('sqExecute').textContent = '⏳ インポート中...';
  try{
    const r = await fetch('/api/import/sqlite/execute', {
      method:'POST',
      headers:{'X-Role':'owner','X-Token':sessionStorage.getItem('pos_token')||''},
      body: fd
    });
    if(!r.ok) throw new Error(await r.text());
    const result = await r.json();
    renderSqDone(result.summary);
    $('sqCard1').classList.add('hidden');
    $('sqCard2').classList.add('hidden');
    $('sqCard3').classList.remove('hidden');
    window.scrollTo({top:0, behavior:'smooth'});
  }catch(e){
    showAlert('error', 'インポート失敗: ' + e.message);
    $('sqExecute').disabled = false;
    $('sqExecute').textContent = '✅ チェックを入れたデータを一括インポート';
  }
});

function renderSqDone(summary){
  let totalIns = 0, totalUpd = 0, totalSkip = 0;
  let rows = '';
  for(const [dt, s] of Object.entries(summary)){
    totalIns += s.inserted || 0;
    totalUpd += s.updated || 0;
    totalSkip += s.skipped || 0;
    rows += `
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:8px;padding:10px 12px;background:#0a1423;border:1px solid var(--line);border-radius:8px;margin-bottom:6px;font-size:13px">
        <div>${TYPE_ICONS[dt]} ${TYPE_LABELS[dt]}</div>
        <div style="color:var(--green)">追加: ${s.inserted||0}</div>
        <div style="color:var(--accent)">更新: ${s.updated||0}</div>
        <div style="color:var(--muted)">スキップ: ${s.skipped||0}</div>
      </div>
    `;
  }
  $('sqCompletion').innerHTML = `
    <div class="alert success">
      🎉 一括インポートが完了しました！
    </div>
    <div style="margin-top:14px">${rows}</div>
    <div style="background:#0a1423;border:1px solid var(--line);border-radius:10px;padding:18px;margin-top:14px;display:grid;grid-template-columns:repeat(3,1fr);gap:14px;text-align:center">
      <div>
        <div style="font-size:32px;font-weight:700;color:var(--green)">${totalIns}</div>
        <div style="font-size:12px;color:var(--muted)">合計新規追加</div>
      </div>
      <div>
        <div style="font-size:32px;font-weight:700;color:var(--accent)">${totalUpd}</div>
        <div style="font-size:12px;color:var(--muted)">合計更新</div>
      </div>
      <div>
        <div style="font-size:32px;font-weight:700;color:var(--muted)">${totalSkip}</div>
        <div style="font-size:12px;color:var(--muted)">合計スキップ</div>
      </div>
    </div>
  `;
}

function resetSqlite(){
  sqFile = null; sqScanData = null;
  $('sqFileInput').value = '';
  $('sqFileName').textContent = '';
  $('sqMatchesArea').innerHTML = '';
  $('sqScanSummary').innerHTML = '';
  $('sqExecute').disabled = false;
  $('sqExecute').textContent = '✅ チェックを入れたデータを一括インポート';
  $('sqCard1').classList.remove('hidden');
  $('sqCard2').classList.add('hidden');
  $('sqCard3').classList.add('hidden');
}
</script>
</body></html>""")
