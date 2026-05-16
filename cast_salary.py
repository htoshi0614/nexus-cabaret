"""cast_salary.py — ドリンクバック・給与計算・Excel書き出し + /ui/salary"""
from datetime import datetime, timezone, date
from typing import Optional, List
from zoneinfo import ZoneInfo
from fastapi import APIRouter, HTTPException, Header, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import Column, Integer, String, Float, Boolean, ForeignKey, DateTime, Text
from sqlalchemy.orm import relationship
import io, json
from db_shared import Base, SessionLocal, require_role

JST = ZoneInfo("Asia/Tokyo")
router = APIRouter(tags=["salary"])
ADMIN_ROLES = ["owner", "manager"]

# ─────────────────────────── DB Models ───────────────────────────

class CastSalaryConfig(Base):
    __tablename__ = "cast_salary_configs"
    id                   = Column(Integer, primary_key=True)
    cast_id              = Column(Integer, ForeignKey("casts.id"), unique=True)
    store_id             = Column(Integer, ForeignKey("stores.id"))
    hourly_rate          = Column(Float, default=0.0)   # 時給
    floor_rate           = Column(Float, default=0.0)   # 場内バック固定額（円/件、0=バックなし）
    drink_back_rate      = Column(Float, default=0.0)   # ドリンクバック率
    bottle_back_rate     = Column(Float, default=0.0)   # ボトルバック率（サービス料抜き小計に対する率）
    nom_fee_hon          = Column(Float, default=0.0)   # 本指名料
    nom_fee_jyonai       = Column(Float, default=0.0)   # 場内指名料
    nom_fee_dohan        = Column(Float, default=0.0)   # 同伴料
    cast                 = relationship("Cast")

class StoreDouhanConfig(Base):
    __tablename__ = "store_douhan_configs"
    id                = Column(Integer, primary_key=True)
    store_id          = Column(Integer, ForeignKey("stores.id"), unique=True)
    douhan_back_rate  = Column(Float, default=0.0)  # 同伴バック率 0-100 (%)
    douhan_fee        = Column(Float, default=0.0)   # 同伴料（固定額、テーブルチャージに追加）

class DrinkBackRecord(Base):
    __tablename__ = "drink_back_records"
    id         = Column(Integer, primary_key=True)
    store_id   = Column(Integer, ForeignKey("stores.id"))
    cast_id    = Column(Integer, ForeignKey("casts.id"))
    session_id = Column(Integer, ForeignKey("sessions.id"))
    order_id   = Column(Integer, ForeignKey("orders.id"))
    back_type  = Column(String, default="drink")  # "drink" or "bottle"
    amount     = Column(Float, default=0.0)   # 実際のバック金額
    created_at = Column(DateTime, default=datetime.utcnow)
    cast       = relationship("Cast")

# ─────────────────────────── Pydantic ───────────────────────────

class CastSalaryConfigIn(BaseModel):
    cast_id: int
    store_id: int
    hourly_rate: float = 0.0
    floor_rate: float = 0.0
    drink_back_rate: float = 0.0
    bottle_back_rate: float = 0.0
    nom_fee_hon: float = 0.0
    nom_fee_jyonai: float = 0.0
    nom_fee_dohan: float = 0.0

class DrinkBackIn(BaseModel):
    cast_id: int
    session_id: int
    order_id: int
    amount: float

# ─────────────────────────── Salary Logic ───────────────────────────

def compute_cast_salary(db, store_id: int, year: int, month: int) -> List[dict]:
    """指定年月のキャスト給与を計算して返す"""
    from sqlalchemy import extract

    # Import models from pos (loaded after pos sets up models)
    try:
        from pos import Cast, Attendance, Nomination
    except ImportError as e:
        import logging
        logging.warning(f"cast_salary: posモデルのインポートに失敗: {e}")
        return []

    casts = db.query(Cast).filter_by(store_id=store_id, is_active=True).all()
    results = []

    for cast in casts:
        cfg = db.query(CastSalaryConfig).filter_by(cast_id=cast.id).first()

        # 勤務時間
        attends = (db.query(Attendance)
                   .filter_by(store_id=store_id, person_type="cast", person_id=cast.id)
                   .all())
        hours_worked = 0.0
        for a in attends:
            if a.clock_in and a.clock_out:
                ci_jst = a.clock_in.replace(tzinfo=timezone.utc).astimezone(JST)
                co_jst = a.clock_out.replace(tzinfo=timezone.utc).astimezone(JST)
                if ci_jst.year == year and ci_jst.month == month:
                    hours_worked += (a.clock_out - a.clock_in).total_seconds() / 3600

        hourly_rate  = cfg.hourly_rate       if cfg else 0.0
        jyonai_back  = cfg.floor_rate         if cfg else 0.0   # 場内指名1件あたり固定バック額（円）
        db_rate      = cfg.drink_back_rate   if cfg else 0.0
        nom_hon      = cfg.nom_fee_hon       if cfg else 0.0
        nom_jyonai   = cfg.nom_fee_jyonai   if cfg else 0.0
        nom_dohan    = cfg.nom_fee_dohan     if cfg else 0.0

        # 時給
        time_pay = hours_worked * hourly_rate

        # ドリンクバック＆ボトルバック
        all_backs = (db.query(DrinkBackRecord)
                     .filter_by(store_id=store_id, cast_id=cast.id)
                     .all())
        drink_back_total = 0.0
        bottle_back_total = 0.0
        for r in all_backs:
            rj = r.created_at.replace(tzinfo=timezone.utc).astimezone(JST)
            if rj.year == year and rj.month == month:
                if getattr(r, 'back_type', 'drink') == 'bottle':
                    bottle_back_total += r.amount
                else:
                    drink_back_total += r.amount

        # 指名料
        noms = (db.query(Nomination).filter_by(store_id=store_id, cast_id=cast.id).all())
        nom_count = {"hon": 0, "jyonai": 0, "dohan": 0}
        nom_pay   = 0.0
        for n in noms:
            created_jst = n.created_at.replace(tzinfo=timezone.utc).astimezone(JST)
            if created_jst.year == year and created_jst.month == month:
                t = n.nomi_type
                if t in nom_count: nom_count[t] += 1
                if t == "hon":    nom_pay += nom_hon
                elif t == "jyonai": nom_pay += nom_jyonai
                elif t == "dohan":  nom_pay += nom_dohan

        # 場内バック（場内指名1件あたり固定額 × 件数、0=バックなし）
        floor_pay = nom_count["jyonai"] * jyonai_back

        total = time_pay + drink_back_total + bottle_back_total + nom_pay + floor_pay

        results.append({
            "cast_id":       cast.id,
            "cast_name":     cast.name,
            "rank":          cast.rank,
            "hours_worked":  round(hours_worked, 2),
            "hourly_rate":   hourly_rate,
            "time_pay":      round(time_pay),
            "drink_back":    round(drink_back_total),
            "bottle_back":   round(bottle_back_total),
            "nom_hon":       nom_count["hon"],
            "nom_jyonai":    nom_count["jyonai"],
            "nom_dohan":     nom_count["dohan"],
            "nom_pay":       round(nom_pay),
            "floor_pay":     round(floor_pay),
            "total":         round(total),
        })

    return results

def export_salary_excel(data: List[dict], year: int, month: int) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d} 給与"

    # ヘッダー
    headers = ["名前", "ランク", "勤務時間(h)", "時給", "時給分",
               "ドリンクバック", "ボトルバック", "本指名", "場内指名", "同伴",
               "指名料計", "場内料", "合計"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    bold = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, d in enumerate(data, 2):
        vals = [d["cast_name"], d["rank"], d["hours_worked"], d["hourly_rate"],
                d["time_pay"], d["drink_back"], d["bottle_back"], d["nom_hon"], d["nom_jyonai"],
                d["nom_dohan"], d["nom_pay"], d["floor_pay"], d["total"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)

    # 通貨フォーマット（¥表示）
    yen_fmt = '#,##0"円"'
    for row_idx in range(2, len(data) + 2):
        for col_idx in [5, 6, 7, 11, 12, 13]:  # 時給分, DB, BB, 指名料, 場内料, 合計
            ws.cell(row=row_idx, column=col_idx).number_format = yen_fmt
        ws.cell(row=row_idx, column=4).number_format = yen_fmt  # 時給

    # 合計行
    last = len(data) + 2
    ws.cell(row=last, column=1, value="合計").font = Font(bold=True)
    for col_idx, key in [(5,"time_pay"),(6,"drink_back"),(7,"bottle_back"),(11,"nom_pay"),(12,"floor_pay"),(13,"total")]:
        total_val = sum(d[key] for d in data)
        cell = ws.cell(row=last, column=col_idx, value=total_val)
        cell.font = Font(bold=True)

    # 列幅調整
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 20)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─────────────────────────── API Routes ───────────────────────────

@router.get("/cast-salary-configs")
def list_salary_configs(store_id: int,
                        x_role: Optional[str] = Header(None, alias="X-Role")):
    require_role(x_role, ADMIN_ROLES)
    db = SessionLocal()
    try:
        try:
            from pos import Cast
        except ImportError as e:
            import logging
            logging.warning(f"cast_salary: Castインポートに失敗: {e}")
            return []
        casts = db.query(Cast).filter_by(store_id=store_id, is_active=True).all()
        result = []
        for c in casts:
            cfg = db.query(CastSalaryConfig).filter_by(cast_id=c.id).first()
            result.append({
                "cast_id": c.id,
                "cast_name": c.name,
                "rank": c.rank,
                "config": {k: v for k, v in cfg.__dict__.items() if not k.startswith("_")} if cfg else None,
            })
        return result
    finally:
        db.close()

@router.post("/cast-salary-configs")
def save_salary_config(payload: CastSalaryConfigIn,
                       x_role: Optional[str] = Header(None, alias="X-Role")):
    require_role(x_role, ADMIN_ROLES)
    db = SessionLocal()
    try:
        cfg = db.query(CastSalaryConfig).filter_by(cast_id=payload.cast_id).first()
        data = payload.model_dump() if hasattr(payload, 'model_dump') else payload.dict()
        if cfg:
            for k, v in data.items():
                setattr(cfg, k, v)
        else:
            cfg = CastSalaryConfig(**data)
            db.add(cfg)
        db.commit()
        return {"ok": True}
    finally:
        db.close()

@router.get("/cast-salary")
def get_salary_report(store_id: int,
                      year: int = Query(default=None),
                      month: int = Query(default=None),
                      x_role: Optional[str] = Header(None, alias="X-Role")):
    require_role(x_role, ADMIN_ROLES)
    now_jst = datetime.now(tz=JST)
    y = year  or now_jst.year
    m = month or now_jst.month
    db = SessionLocal()
    try:
        data = compute_cast_salary(db, store_id, y, m)
        return {"year": y, "month": m, "store_id": store_id, "casts": data}
    finally:
        db.close()

@router.get("/cast-salary/export")
def export_salary(store_id: int,
                  year: int = Query(default=None),
                  month: int = Query(default=None),
                  x_role: Optional[str] = Header(None, alias="X-Role")):
    require_role(x_role, ADMIN_ROLES)
    now_jst = datetime.now(tz=JST)
    y = year  or now_jst.year
    m = month or now_jst.month
    db = SessionLocal()
    try:
        data = compute_cast_salary(db, store_id, y, m)
        xlsx = export_salary_excel(data, y, m)
        filename = f"salary_{y}{m:02d}.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        db.close()

@router.post("/drink-back")
def record_drink_back(store_id: int, payload: DrinkBackIn,
                      x_role: Optional[str] = Header(None, alias="X-Role")):
    require_role(x_role, ADMIN_ROLES)
    db = SessionLocal()
    try:
        rec = DrinkBackRecord(store_id=store_id, **(payload.model_dump() if hasattr(payload, 'model_dump') else payload.dict()))
        db.add(rec); db.commit()
        return {"ok": True}
    finally:
        db.close()

# ─────────────────────────── 同伴バック設定 API ───────────────────────────

@router.get("/douhan-config")
def get_douhan_config(store_id: int, x_role: Optional[str] = Header(None, alias="X-Role")):
    require_role(x_role, ["owner","manager","cashier","staff"])
    db = SessionLocal()
    try:
        cfg = db.query(StoreDouhanConfig).filter_by(store_id=store_id).first()
        if not cfg:
            return {"store_id": store_id, "douhan_back_rate": 0.0, "douhan_fee": 0.0}
        return {"store_id": cfg.store_id, "douhan_back_rate": cfg.douhan_back_rate, "douhan_fee": cfg.douhan_fee}
    finally:
        db.close()

@router.post("/douhan-config")
def save_douhan_config(payload: dict, x_role: Optional[str] = Header(None, alias="X-Role")):
    require_role(x_role, ["owner","manager"])
    store_id = payload.get("store_id")
    rate = payload.get("douhan_back_rate", 0.0)
    fee = payload.get("douhan_fee", 0.0)
    if rate < 0 or rate > 100:
        raise HTTPException(400, "douhan_back_rate must be 0-100")
    db = SessionLocal()
    try:
        cfg = db.query(StoreDouhanConfig).filter_by(store_id=store_id).first()
        if cfg:
            cfg.douhan_back_rate = rate
            cfg.douhan_fee = fee
        else:
            cfg = StoreDouhanConfig(store_id=store_id, douhan_back_rate=rate, douhan_fee=fee)
            db.add(cfg)
        db.commit()
        return {"ok": True}
    finally:
        db.close()

@router.get("/douhan-stats")
def get_douhan_stats(store_id: int, year: int = None, month: int = None,
                     x_role: Optional[str] = Header(None, alias="X-Role")):
    """同伴数・同伴率の集計"""
    require_role(x_role, ADMIN_ROLES)
    now_jst = datetime.now(tz=JST)
    y = year or now_jst.year
    m = month or now_jst.month
    db = SessionLocal()
    try:
        from pos import Cast, Nomination, Session
        casts = db.query(Cast).filter_by(store_id=store_id, is_active=True).all()
        total_sessions = db.query(Session).filter_by(store_id=store_id).count()
        results = []
        for c in casts:
            noms = db.query(Nomination).filter_by(store_id=store_id, cast_id=c.id, nomi_type="dohan").all()
            month_count = 0
            total_count = 0
            for n in noms:
                total_count += 1
                nj = n.created_at.replace(tzinfo=timezone.utc).astimezone(JST)
                if nj.year == y and nj.month == m:
                    month_count += 1
            # 同伴率 = 月間同伴数 / 月間出勤日数
            from pos import Attendance
            attends = db.query(Attendance).filter_by(store_id=store_id, person_type="cast", person_id=c.id).all()
            work_days = 0
            for a in attends:
                if a.clock_in:
                    aj = a.clock_in.replace(tzinfo=timezone.utc).astimezone(JST)
                    if aj.year == y and aj.month == m:
                        work_days += 1
            douhan_rate = (month_count / work_days * 100) if work_days > 0 else 0
            results.append({
                "cast_id": c.id, "cast_name": c.name,
                "month_count": month_count, "total_count": total_count,
                "work_days": work_days, "douhan_rate": round(douhan_rate, 1)
            })
        return {"year": y, "month": m, "casts": results}
    finally:
        db.close()

# ─────────────────────────── Salary UI ───────────────────────────

@router.get("/ui/salary", response_class=HTMLResponse)
def ui_salary():
    return HTMLResponse(r"""
<!doctype html><html lang="ja"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>給与管理 - Girls Bar POS</title>
<style>
:root{--bg:#0b1220;--card:#0f172a;--line:#1f2937;--text:#e5e7eb;--muted:#94a3b8;--accent:#0ea5e9}
*{box-sizing:border-box;font-family:-apple-system,system-ui,"Noto Sans JP",sans-serif}
body{margin:0;background:var(--bg);color:var(--text)}
header{position:sticky;top:0;z-index:40;display:flex;gap:12px;align-items:center;padding:12px 16px;border-bottom:1px solid var(--line);background:rgba(11,18,32,.95);backdrop-filter:blur(6px)}
header h1{margin:0;font-size:17px}
.nav a{color:var(--accent);text-decoration:none;font-size:14px;padding:6px 10px;border-radius:8px;border:1px solid var(--line)}
.container{max-width:1100px;margin:0 auto;padding:20px 16px;display:flex;flex-direction:column;gap:20px}
.card{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:20px}
.card h2{margin:0 0 14px;font-size:15px;border-bottom:1px solid var(--line);padding-bottom:10px}
label{display:flex;flex-direction:column;gap:4px;font-size:13px;color:var(--muted)}
input,select{font-size:14px;padding:7px 10px;border-radius:8px;border:1px solid #263244;background:#0a1220;color:var(--text)}
.btn{cursor:pointer;font-size:14px;padding:8px 16px;border-radius:10px;border:1px solid #334155;background:#111827;color:var(--text)}
.btn.solid{background:var(--accent);border-color:var(--accent);color:#001018;font-weight:700}
.btn.green{background:#14532d;border-color:#22c55e;color:#86efac}
.row{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
table{width:100%;border-collapse:collapse;font-size:13px}
th,td{padding:8px 10px;border-bottom:1px solid var(--line);text-align:left}
th{color:var(--muted);font-weight:500}
td.num{text-align:right;font-family:monospace}
.edit-form{background:#0a1624;border:1px solid var(--line);border-radius:12px;padding:14px;margin-top:12px;display:none}
.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
@media(max-width:700px){
  .grid3{grid-template-columns:1fr}
  .container{padding:12px 10px}
  table{font-size:11px;display:block;overflow-x:auto}
  th,td{padding:6px 4px;white-space:nowrap}
  header{flex-wrap:wrap;gap:8px}
  .row{flex-direction:column;align-items:stretch}
}
</style></head><body>
<header>
  <h1>給与管理</h1>
  <div class="nav" style="display:flex;gap:8px;margin-left:auto">
    <a href="/ui">← フロア</a>
    <a href="/ui/pricing">料金設定</a>
    <a href="/ui/weather">天気/シフト</a>
    <a href="/ui/subscription">サブスク</a>
  </div>
</header>

<div class="container">
  <div class="row">
    <label style="flex-direction:row;align-items:center;gap:6px">店舗 <input id="storeId" type="number" value="1" style="width:70px"></label>
    <label style="flex-direction:row;align-items:center;gap:6px">年 <input id="year" type="number" style="width:90px"></label>
    <label style="flex-direction:row;align-items:center;gap:6px">月 <input id="month" type="number" min="1" max="12" style="width:60px"></label>
    <button class="btn solid" onclick="loadReport()">給与計算</button>
    <button class="btn green" onclick="exportExcel()">Excel書き出し</button>
    <button class="btn" onclick="loadConfigs()">給与設定を表示</button>
    <button class="btn" onclick="loadDouhanStats()" style="background:#4a1942;border-color:#e879f9;color:#f0abfc">同伴集計</button>
    <button class="btn" onclick="toggleDouhanConfig()" style="background:#3b0764;border-color:#a855f7;color:#c084fc">同伴バック率設定</button>
  </div>

  <!-- 同伴バック率設定（店舗単位） -->
  <div class="card" id="douhanConfigCard" style="display:none">
    <h2>同伴バック率設定（店舗共通）</h2>
    <p style="color:var(--muted);font-size:12px;margin-bottom:12px">マネージャー以上のみ変更可能</p>
    <div class="grid3">
      <label>同伴バック率 (0-100%)
        <div class="row" style="gap:6px;align-items:center">
          <input id="douhanRate" type="number" min="0" max="100" step="1" style="width:100px">
          <span>%</span>
        </div>
      </label>
      <label>同伴料（固定額/回）
        <div class="row" style="gap:6px;align-items:center">
          <span>¥</span><input id="douhanFee" type="number" min="0" step="100">
        </div>
      </label>
    </div>
    <div class="row" style="margin-top:12px;justify-content:flex-end">
      <button class="btn solid" onclick="saveDouhanConfig()">保存</button>
    </div>
  </div>

  <!-- 同伴集計 -->
  <div class="card" id="douhanStatsCard" style="display:none">
    <h2 id="douhanStatsTitle">同伴集計</h2>
    <table>
      <thead><tr><th>名前</th><th>月間同伴数</th><th>累計同伴数</th><th>月間出勤日</th><th>同伴率</th></tr></thead>
      <tbody id="douhanStatsBody"></tbody>
    </table>
  </div>

  <!-- 給与レポート -->
  <div class="card" id="reportCard" style="display:none">
    <h2 id="reportTitle">給与レポート</h2>
    <table>
      <thead><tr>
        <th>名前</th><th>ランク</th><th>勤務時間</th><th>時給分</th>
        <th>ドリンクバック</th><th>ボトルバック</th><th>本指名</th><th>場内指名</th><th>同伴</th>
        <th>指名料</th><th>場内料</th><th style="text-align:right">合計</th>
      </tr></thead>
      <tbody id="reportBody"></tbody>
      <tfoot id="reportFoot"></tfoot>
    </table>
  </div>

  <!-- 給与設定 -->
  <div class="card" id="configCard" style="display:none">
    <h2>給与設定（キャスト別）</h2>
    <table>
      <thead><tr>
        <th>名前</th><th>時給</th><th>場内バック/件</th><th>ドリンクバック率</th><th>ボトルバック率</th>
        <th>本指名</th><th>場内指名</th><th>同伴</th><th></th>
      </tr></thead>
      <tbody id="configBody"></tbody>
    </table>
  </div>
</div>

<script>
const $ = id=>document.getElementById(id);
const now = new Date();
$('year').value = now.getFullYear();
$('month').value = now.getMonth()+1;

async function api(path,opt={}){
  const tk=sessionStorage.getItem('pos_token')||'';
  const o={method:'GET',headers:{'Content-Type':'application/json','X-Role':'owner','X-Token':tk},...opt};
  if(o.body&&typeof o.body!=='string') o.body=JSON.stringify(o.body);
  const r=await fetch(path,o);
  if(r.status===401){sessionStorage.clear();window.location.href='/';return;}
  if(!r.ok) throw new Error(await r.text());
  return r.json();
}

async function loadReport(){
  const s=$('storeId').value, y=$('year').value, m=$('month').value;
  try{
    const d=await api(`/cast-salary?store_id=${s}&year=${y}&month=${m}`);
    $('reportCard').style.display='';
    $('reportTitle').textContent=`${y}年${m}月 給与レポート`;
    const tb=$('reportBody');
    tb.innerHTML='';
    let totTime=0,totDB=0,totBB=0,totNom=0,totFloor=0,totAll=0;
    (d.casts||[]).forEach(c=>{
      const tr=document.createElement('tr');
      tr.innerHTML=`<td>${c.cast_name}</td><td>${c.rank||''}</td>
        <td class="num">${c.hours_worked}h</td><td class="num">¥${(c.time_pay||0).toLocaleString()}</td>
        <td class="num">¥${(c.drink_back||0).toLocaleString()}</td>
        <td class="num">¥${(c.bottle_back||0).toLocaleString()}</td>
        <td class="num">${c.nom_hon}</td><td class="num">${c.nom_jyonai}</td><td class="num">${c.nom_dohan}</td>
        <td class="num">¥${(c.nom_pay||0).toLocaleString()}</td>
        <td class="num">¥${(c.floor_pay||0).toLocaleString()}</td>
        <td class="num"><b>¥${(c.total||0).toLocaleString()}</b></td>`;
      tb.appendChild(tr);
      totTime+=c.time_pay||0; totDB+=c.drink_back||0; totBB+=c.bottle_back||0;
      totNom+=c.nom_pay||0; totFloor+=c.floor_pay||0; totAll+=c.total||0;
    });
    $('reportFoot').innerHTML=`<tr style="border-top:2px solid var(--accent)">
      <td colspan="3"><b>合計</b></td>
      <td class="num">¥${totTime.toLocaleString()}</td>
      <td class="num">¥${totDB.toLocaleString()}</td>
      <td class="num">¥${totBB.toLocaleString()}</td>
      <td colspan="3"></td>
      <td class="num">¥${totNom.toLocaleString()}</td>
      <td class="num">¥${totFloor.toLocaleString()}</td>
      <td class="num"><b>¥${totAll.toLocaleString()}</b></td></tr>`;
  }catch(e){alert(e.message)}
}

async function exportExcel(){
  const s=$('storeId').value, y=$('year').value, m=$('month').value;
  window.location.href=`/cast-salary/export?store_id=${s}&year=${y}&month=${m}`;
}

async function loadConfigs(){
  const s=$('storeId').value;
  try{
    const data=await api(`/cast-salary-configs?store_id=${s}`);
    $('configCard').style.display='';
    const tb=$('configBody'); tb.innerHTML='';
    data.forEach(d=>{
      const c=d.config||{};
      const tr=document.createElement('tr');
      tr.innerHTML=`<td><b>${d.cast_name}</b></td>
        <td class="num">¥${(c.hourly_rate||0).toLocaleString()}</td>
        <td class="num">¥${(c.floor_rate||0).toLocaleString()}</td>
        <td class="num">${((c.drink_back_rate||0)*100).toFixed(0)}%</td>
        <td class="num">${((c.bottle_back_rate||0)*100).toFixed(0)}%</td>
        <td class="num">¥${(c.nom_fee_hon||0).toLocaleString()}</td>
        <td class="num">¥${(c.nom_fee_jyonai||0).toLocaleString()}</td>
        <td class="num">¥${(c.nom_fee_dohan||0).toLocaleString()}</td>
        <td><button class="btn" onclick="openEdit(${d.cast_id},${JSON.stringify(c).replace(/"/g,'&quot;')})">編集</button></td>`;
      tb.appendChild(tr);
    });
  }catch(e){alert(e.message)}
}

function openEdit(castId, cfg){
  const existing = document.getElementById('editForm');
  if(existing) existing.remove();
  const store = $('storeId').value;
  const div=document.createElement('div');
  div.id='editForm'; div.className='edit-form'; div.style.display='block';
  div.innerHTML=`<div class="grid3">
    <label>時給 <input id="e_hr" type="number" value="${cfg.hourly_rate||0}"></label>
    <label>場内バック（円/件、0=バックなし）<input id="e_fr" type="number" step="1" min="0" value="${cfg.floor_rate||0}"></label>
    <label>ドリンクバック率 <input id="e_db" type="number" step="0.01" value="${cfg.drink_back_rate||0}"></label>
    <label>ボトルバック率 <input id="e_bb" type="number" step="0.01" value="${cfg.bottle_back_rate||0}"></label>
    <label>本指名料 <input id="e_nh" type="number" value="${cfg.nom_fee_hon||0}"></label>
    <label>場内指名料 <input id="e_nj" type="number" value="${cfg.nom_fee_jyonai||0}"></label>
    <label>同伴料 <input id="e_nd" type="number" value="${cfg.nom_fee_dohan||0}"></label>
  </div>
  <div class="row" style="margin-top:12px;justify-content:flex-end">
    <button class="btn solid" onclick="saveConfig(${castId},${store})">保存</button>
    <button class="btn" onclick="document.getElementById('editForm').remove()">キャンセル</button>
  </div>`;
  $('configCard').appendChild(div);
}

async function saveConfig(castId, storeId){
  try{
    await api('/cast-salary-configs',{method:'POST',body:{
      cast_id:castId, store_id:storeId,
      hourly_rate:parseFloat($('e_hr').value||0),
      floor_rate:parseFloat($('e_fr').value||0),
      drink_back_rate:parseFloat($('e_db').value||0),
      bottle_back_rate:parseFloat($('e_bb').value||0),
      nom_fee_hon:parseFloat($('e_nh').value||0),
      nom_fee_jyonai:parseFloat($('e_nj').value||0),
      nom_fee_dohan:parseFloat($('e_nd').value||0),
    }});
    alert('保存しました'); document.getElementById('editForm')?.remove();
    loadConfigs();
  }catch(e){alert(e.message)}
}

/* ====== 同伴バック率設定 ====== */
async function toggleDouhanConfig(){
  const card=$('douhanConfigCard');
  if(card.style.display==='none'){
    card.style.display='';
    const s=$('storeId').value;
    try{
      const d=await api(`/douhan-config?store_id=${s}`);
      $('douhanRate').value=d.douhan_back_rate||0;
      $('douhanFee').value=d.douhan_fee||0;
    }catch{}
  }else{
    card.style.display='none';
  }
}
async function saveDouhanConfig(){
  const s=$('storeId').value;
  const rate=parseFloat($('douhanRate').value||0);
  const fee=parseFloat($('douhanFee').value||0);
  if(rate<0||rate>100) return alert('バック率は0-100で入力してください');
  try{
    await api('/douhan-config',{method:'POST',body:{store_id:parseInt(s),douhan_back_rate:rate,douhan_fee:fee}});
    alert('同伴バック率を保存しました');
  }catch(e){alert(e.message)}
}

/* ====== 同伴集計 ====== */
async function loadDouhanStats(){
  const s=$('storeId').value, y=$('year').value, m=$('month').value;
  try{
    const d=await api(`/douhan-stats?store_id=${s}&year=${y}&month=${m}`);
    $('douhanStatsCard').style.display='';
    $('douhanStatsTitle').textContent=`${y}年${m}月 同伴集計`;
    const tb=$('douhanStatsBody'); tb.innerHTML='';
    (d.casts||[]).forEach(c=>{
      const tr=document.createElement('tr');
      tr.innerHTML=`<td>${c.cast_name}</td>
        <td class="num">${c.month_count}</td>
        <td class="num">${c.total_count}</td>
        <td class="num">${c.work_days}</td>
        <td class="num" style="color:${c.douhan_rate>=50?'#22c55e':c.douhan_rate>=30?'#facc15':'#ef4444'};font-weight:700">${c.douhan_rate}%</td>`;
      tb.appendChild(tr);
    });
  }catch(e){alert(e.message)}
}
</script>
</body></html>
""")
